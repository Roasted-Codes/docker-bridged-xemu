#!/usr/bin/env python3
"""
Export Halo 2 game history to Excel (.xlsx).

Modes:
  Default (all-in-one): One workbook with Games, Players, Player Summary, Teams sheets.
  --per-game:           One workbook per game with detailed sheets.

Per-game styles:
  --style bungie   (default)  Clean Halo 2 themed presentation. Only real data.
  --style rampant             Exact column match with Rampant's spreadsheet format.

Requires: openpyxl (pip install openpyxl)
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required for Excel export.", file=sys.stderr)
    print("Install it with:  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ===========================================================================
# Styling constants — all-in-one mode
# ===========================================================================
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center")
COL_WIDTH_PADDING = 4

# ===========================================================================
# Styling constants — Bungie mode
# ===========================================================================
BUNGIE_DARK = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
BUNGIE_HEADER_FONT = Font(bold=True, color="c4a747", size=11)
BUNGIE_TITLE_FONT = Font(bold=True, size=14, color="c4a747")
BUNGIE_SUBTITLE_FONT = Font(bold=True, size=11, color="888888")
BUNGIE_LABEL_FONT = Font(bold=True, size=11, color="AAAAAA")
BUNGIE_VALUE_FONT = Font(size=11, color="FFFFFF")
BUNGIE_ACCENT = PatternFill(start_color="2a2a3e", end_color="2a2a3e", fill_type="solid")
BUNGIE_GOLD_ROW = PatternFill(start_color="3d3520", end_color="3d3520", fill_type="solid")
BUNGIE_GRAY_DIAG = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
BUNGIE_THIN_BORDER = Border(
    left=Side(style="thin", color="333333"),
    right=Side(style="thin", color="333333"),
    top=Side(style="thin", color="333333"),
    bottom=Side(style="thin", color="333333"),
)

# ===========================================================================
# Regex to strip characters illegal in XLSX cells
# ===========================================================================
_ILLEGAL_XML_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f"
    r"\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]"
)

# ===========================================================================
# Medal and gametype constants
# ===========================================================================
MEDAL_COLUMNS = [
    "double_kill", "triple_kill", "killtacular", "kill_frenzy",
    "killtrocity", "killamanjaro", "sniper_kill", "road_kill",
    "bone_cracker", "assassin", "vehicle_destroyed", "car_jacking",
    "stick_it", "killing_spree", "running_riot", "rampage",
    "berserker", "over_kill", "flag_taken", "flag_carrier_kill",
    "flag_returned", "bomb_planted", "bomb_carrier_kill", "bomb_returned",
]

MEDAL_DISPLAY_NAMES = [
    "Double Kill", "Triple Kill", "Killtacular", "Kill Frenzy",
    "Killtrocity", "Killimanjaro", "Sniper Kill", "Splatter",
    "Beat Down", "Assassination", "Destroyed Vehicle", "Carjack",
    "Stick", "Killing Spree", "Running Riot", "Rampage",
    "Berserker", "Overkill", "Flag Grab", "Flag Carrier Kill",
    "Flag Returned", "Bomb Planted", "Bomb Carrier Kill", "Bomb Defused",
]

GAMETYPE_STAT_COLUMNS = [
    "ctf_scores", "ctf_flag_steals", "ctf_flag_saves", "ctf_unknown",
    "assault_score", "assault_bomber_kills", "assault_bomb_grabbed",
    "oddball_score", "oddball_ball_kills", "oddball_carried_kills",
    "koth_kills_as_king", "koth_kings_killed",
    "juggernauts_killed", "kills_as_juggernaut", "juggernaut_time",
    "territories_taken", "territories_lost",
]

# Maps gametype name -> (value0_column, value1_column) in GAMETYPE_STAT_COLUMNS
GAMETYPE_VALUE_MAP = {
    "ctf":          ("ctf_flag_saves", "ctf_flag_steals"),
    "slayer":       ("total_time_alive", "best_spree"),
    "oddball":      ("oddball_ball_kills", "oddball_carried_kills"),
    "koth":         ("koth_kills_as_king", "koth_kings_killed"),
    "juggernaut":   ("juggernauts_killed", "kills_as_juggernaut"),
    "territories":  ("territories_taken", "territories_lost"),
    "assault":      ("assault_bomb_grabbed", "assault_bomber_kills"),
}

# Gametype-specific label pairs for Bungie mode
GAMETYPE_LABELS = {
    "ctf":          ("Flag Saves", "Flag Steals"),
    "slayer":       ("Avg Life", "Best Spree"),
    "oddball":      ("Carrier Kills", "Ball Kills"),
    "koth":         ("Kings Killed", "Kills From"),
    "juggernaut":   ("Jugs Killed", "Kills As Jug"),
    "territories":  ("Terr. Taken", "Terr. Lost"),
    "assault":      ("Bomb Grabs", "Bomber Kills"),
}


# ===========================================================================
# Shared helpers
# ===========================================================================

def _sanitize(value):
    """Strip illegal XML characters from strings before writing to Excel."""
    if isinstance(value, str):
        return _ILLEGAL_XML_RE.sub("", value)
    return value


def _safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _player_accuracy(player: dict) -> float:
    acc = player.get("accuracy")
    if isinstance(acc, dict):
        return _safe_float(acc.get("percentage"))
    return 0.0


def _player_headshots(player: dict) -> int:
    acc = player.get("accuracy")
    if isinstance(acc, dict):
        return _safe_int(acc.get("headshots"))
    return 0


def _player_total_shots(player: dict) -> int:
    acc = player.get("accuracy")
    if isinstance(acc, dict):
        return _safe_int(acc.get("total_shots"))
    return 0


def _player_shots_hit(player: dict) -> int:
    acc = player.get("accuracy")
    if isinstance(acc, dict):
        return _safe_int(acc.get("shots_hit"))
    return 0


def _player_medals_total(player: dict) -> int:
    medals = player.get("medals")
    if isinstance(medals, dict):
        return _safe_int(medals.get("total"))
    return 0


def _player_medals_bitmask(player: dict) -> int:
    medals = player.get("medals")
    if isinstance(medals, dict):
        return _safe_int(medals.get("by_type"))
    return 0


def _player_kd(player: dict) -> float:
    kills = _safe_int(player.get("kills"))
    deaths = _safe_int(player.get("deaths"))
    if deaths == 0:
        return float(kills)
    return kills / deaths


def _player_kda(player: dict) -> float:
    kills = _safe_int(player.get("kills"))
    assists = _safe_int(player.get("assists"))
    deaths = _safe_int(player.get("deaths"))
    return (kills + assists) / max(deaths, 1)


def _gt_values(player: dict):
    vals = player.get("gametype_values")
    if isinstance(vals, list) and len(vals) >= 2:
        return _safe_int(vals[0]), _safe_int(vals[1])
    return 0, 0


def _get_killed_array(player: dict) -> list:
    """Get killed array, handling both 'killed' and 'killed_by' field names."""
    return player.get("killed") or player.get("killed_by") or [0] * 16


def _decode_medal_bits(by_type: int) -> list:
    """Decode 24-bit medal bitmask to list of 24 ints (0 or 1)."""
    return [(by_type >> i) & 1 for i in range(24)]


def _is_valid_name(name: str) -> bool:
    """Check if a player name is valid printable ASCII."""
    if not name or not name.strip():
        return False
    return all(0x20 <= ord(c) <= 0x7E for c in name)


def _get_valid_players(game: dict) -> list:
    """Return only players with valid ASCII names."""
    return [p for p in game.get("players", []) if _is_valid_name(p.get("name", ""))]


def _game_has_valid_players(game: dict) -> bool:
    """Check if game has at least one player with a valid name."""
    return len(_get_valid_players(game)) > 0


def _append_row(ws, row):
    """Append a row, sanitizing all string values."""
    ws.append([_sanitize(v) for v in row])


def _style_header_row(ws):
    """Apply bold/dark styling to the first row and freeze it."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    ws.freeze_panes = "A2"


def _auto_column_widths(ws, headers):
    """Set column widths based on header length plus padding."""
    for idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = len(str(header)) + COL_WIDTH_PADDING


def _load_games(history_dir: str, date_from: str = None, date_to: str = None):
    """Load and filter JSON game files from history_dir, sorted by timestamp."""
    history_path = Path(history_dir)
    if not history_path.is_dir():
        print(f"Error: History directory not found: {history_dir}", file=sys.stderr)
        sys.exit(1)

    games = []
    for fpath in sorted(history_path.glob("*.json")):
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as exc:
            print(f"Warning: skipping {fpath.name}: {exc}", file=sys.stderr)
            continue

        ts_str = data.get("timestamp", "")
        try:
            ts = datetime.fromisoformat(ts_str)
        except (ValueError, TypeError):
            print(f"Warning: skipping {fpath.name}: bad timestamp", file=sys.stderr)
            continue

        if date_from:
            try:
                dt_from = datetime.strptime(date_from, "%Y-%m-%d")
                if ts < dt_from:
                    continue
            except ValueError:
                pass
        if date_to:
            try:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(
                    hour=23, minute=59, second=59
                )
                if ts > dt_to:
                    continue
            except ValueError:
                pass

        data["_timestamp"] = ts
        data["_filename"] = fpath.name
        games.append(data)

    games.sort(key=lambda g: g["_timestamp"])
    return games


# ===========================================================================
# All-in-one sheet builders (existing mode)
# ===========================================================================

def _build_games_sheet(wb, games):
    ws = wb.active
    ws.title = "Games"
    headers = ["Game #", "Date", "Time", "Gametype", "Player Count", "Source", "Fingerprint"]
    ws.append(headers)
    for game_num, g in enumerate(games, 1):
        ts = g["_timestamp"]
        g["_game_num"] = game_num
        _append_row(ws, [
            game_num, ts.strftime("%Y-%m-%d"), ts.strftime("%H:%M:%S"),
            g.get("gametype") or "", _safe_int(g.get("player_count")),
            g.get("source", ""), g.get("fingerprint", ""),
        ])
    _style_header_row(ws)
    _auto_column_widths(ws, headers)


def _build_players_sheet(wb, games):
    ws = wb.create_sheet("Players")
    headers = [
        "Game #", "Date", "Time", "Gametype", "Player", "Place", "Score",
        "Kills", "Deaths", "Assists", "Suicides", "K/D",
        "Accuracy%", "Headshots", "Medals", "GT Value 1", "GT Value 2",
    ]
    ws.append(headers)
    kd_col = headers.index("K/D") + 1
    acc_col = headers.index("Accuracy%") + 1
    for g in games:
        game_num = g.get("_game_num", 0)
        ts = g["_timestamp"]
        date_str = ts.strftime("%Y-%m-%d")
        time_str = ts.strftime("%H:%M:%S")
        gametype = g.get("gametype") or ""
        sorted_players = sorted(
            g.get("players", []),
            key=lambda p: _safe_int(p.get("kills")), reverse=True
        )
        for p in sorted_players:
            gt0, gt1 = _gt_values(p)
            _append_row(ws, [
                game_num, date_str, time_str, gametype,
                p.get("name", ""), _safe_int(p.get("place")),
                p.get("score_string", ""), _safe_int(p.get("kills")),
                _safe_int(p.get("deaths")), _safe_int(p.get("assists")),
                _safe_int(p.get("suicides")), round(_player_kd(p), 2),
                round(_player_accuracy(p), 1), _player_headshots(p),
                _player_medals_total(p), gt0, gt1,
            ])
    for row in ws.iter_rows(min_row=2, min_col=kd_col, max_col=kd_col):
        for cell in row:
            cell.number_format = "0.00"
    for row in ws.iter_rows(min_row=2, min_col=acc_col, max_col=acc_col):
        for cell in row:
            cell.number_format = "0.0"
    _style_header_row(ws)
    _auto_column_widths(ws, headers)


def _build_summary_sheet(wb, games):
    ws = wb.create_sheet("Player Summary")
    headers = [
        "Player", "Games", "Total Kills", "Total Deaths", "Total Assists",
        "Avg K/D", "Best Game K/D", "Avg Accuracy%",
        "Total Headshots", "Total Medals",
    ]
    ws.append(headers)
    agg = {}
    for g in games:
        for p in g.get("players", []):
            name = p.get("name", "")
            if not name:
                continue
            if name not in agg:
                agg[name] = {
                    "games": 0, "kills": 0, "deaths": 0, "assists": 0,
                    "total_shots": 0, "shots_hit": 0, "headshots": 0,
                    "medals": 0, "best_kd": 0.0,
                }
            s = agg[name]
            s["games"] += 1
            s["kills"] += _safe_int(p.get("kills"))
            s["deaths"] += _safe_int(p.get("deaths"))
            s["assists"] += _safe_int(p.get("assists"))
            s["total_shots"] += _player_total_shots(p)
            s["shots_hit"] += _player_shots_hit(p)
            s["headshots"] += _player_headshots(p)
            s["medals"] += _player_medals_total(p)
            game_kd = _player_kd(p)
            if game_kd > s["best_kd"]:
                s["best_kd"] = game_kd
    sorted_players = sorted(agg.items(), key=lambda kv: kv[1]["kills"], reverse=True)
    for name, s in sorted_players:
        avg_kd = s["kills"] / max(s["deaths"], 1)
        avg_acc = (s["shots_hit"] / s["total_shots"] * 100) if s["total_shots"] > 0 else 0.0
        _append_row(ws, [
            name, s["games"], s["kills"], s["deaths"], s["assists"],
            round(avg_kd, 2), round(s["best_kd"], 2), round(avg_acc, 1),
            s["headshots"], s["medals"],
        ])
    _style_header_row(ws)
    _auto_column_widths(ws, headers)


def _build_teams_sheet(wb, games):
    ws = wb.create_sheet("Teams")
    headers = ["Game #", "Date", "Time", "Gametype", "Team", "Score", "Place"]
    ws.append(headers)
    for g in games:
        teams = g.get("teams")
        if not teams:
            continue
        game_num = g.get("_game_num", 0)
        ts = g["_timestamp"]
        for t in teams:
            _append_row(ws, [
                game_num, ts.strftime("%Y-%m-%d"), ts.strftime("%H:%M:%S"),
                g.get("gametype") or "", t.get("name", ""),
                _safe_int(t.get("score")), _safe_int(t.get("place")),
            ])
    _style_header_row(ws)
    _auto_column_widths(ws, headers)


# ===========================================================================
# Per-game: RAMPANT mode
# ===========================================================================

def _rampant_game_details(wb, game):
    """Sheet 1: Game Details — key-value pairs."""
    ws = wb.active
    ws.title = "Game Details"
    ts = game["_timestamp"]
    gametype = game.get("gametype") or ""
    rows = [
        ("Game Type", gametype.capitalize() if gametype else ""),
        ("Variant Name", ""),
        ("Map Name", ""),
        ("Start Time", ts.strftime("%Y-%m-%d %H:%M:%S")),
        ("End Time", ""),
        ("Duration", ""),
    ]
    for label, value in rows:
        _append_row(ws, [label, value])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    # Style labels bold
    for row_num in range(1, len(rows) + 1):
        ws.cell(row=row_num, column=1).font = Font(bold=True)


def _rampant_post_game_report(wb, game):
    """Sheet 2: Post Game Report — one row per player."""
    ws = wb.create_sheet("Post Game Report")
    headers = [
        "name", "place", "score", "kills", "deaths", "assists",
        "kda", "suicides", "team", "shots_fired", "shots_hit",
        "accuracy", "head_shots",
    ]
    ws.append(headers)
    players = _get_valid_players(game)
    sorted_players = sorted(players, key=lambda p: _safe_int(p.get("place")))
    for p in sorted_players:
        _append_row(ws, [
            p.get("name", ""),
            _safe_int(p.get("place")) + 1,
            _sanitize(p.get("score_string", "")),
            _safe_int(p.get("kills")),
            _safe_int(p.get("deaths")),
            _safe_int(p.get("assists")),
            round(_player_kda(p), 2),
            _safe_int(p.get("suicides")),
            "",  # team — player→team mapping not in JSON
            _player_total_shots(p),
            _player_shots_hit(p),
            round(_player_accuracy(p), 1),
            _player_headshots(p),
        ])
    _style_header_row(ws)
    _auto_column_widths(ws, headers)


def _rampant_versus(wb, game):
    """Sheet 3: Versus — NxN player kill matrix."""
    ws = wb.create_sheet("Versus")
    players = _get_valid_players(game)
    if not players:
        ws.append(["X"])
        return
    names = [p.get("name", "") for p in players]
    # Header row: "X" + player names
    ws.append(["X"] + [_sanitize(n) for n in names])
    for i, p in enumerate(players):
        killed = _get_killed_array(p)
        row = [_sanitize(names[i])]
        for j in range(len(players)):
            # killed[j] = how many times player i killed player slot j
            # Need slot index mapping — players are stored in slot order
            row.append(_safe_int(killed[j]) if j < len(killed) else 0)
        ws.append(row)
    # Style header
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row_num in range(2, len(players) + 2):
        ws.cell(row=row_num, column=1).font = Font(bold=True)
    _auto_column_widths(ws, ["X"] + names)


def _rampant_game_statistics(wb, game):
    """Sheet 4: Game Statistics — extended stats with all gametype columns."""
    ws = wb.create_sheet("Game Statistics")
    headers = [
        "Player", "Emblem URL", "kills", "assists", "deaths", "headshots",
        "betrayals", "suicides", "best_spree", "total_time_alive",
    ] + GAMETYPE_STAT_COLUMNS
    ws.append(headers)

    gametype = (game.get("gametype") or "").lower()
    players = _get_valid_players(game)
    sorted_players = sorted(players, key=lambda p: _safe_int(p.get("place")))

    for p in sorted_players:
        gt0, gt1 = _gt_values(p)
        # Build gametype column values — all zeros except the relevant pair
        gt_cols = {col: 0 for col in GAMETYPE_STAT_COLUMNS}
        if gametype in GAMETYPE_VALUE_MAP:
            col0, col1 = GAMETYPE_VALUE_MAP[gametype]
            if col0 in gt_cols:
                gt_cols[col0] = gt0
            if col1 in gt_cols:
                gt_cols[col1] = gt1

        # best_spree and total_time_alive are Slayer-specific via gametype_values
        best_spree = gt1 if gametype == "slayer" else ""
        total_time_alive = gt0 if gametype == "slayer" else ""

        row = [
            p.get("name", ""),
            "",  # Emblem URL — not available
            _safe_int(p.get("kills")),
            _safe_int(p.get("assists")),
            _safe_int(p.get("deaths")),
            _player_headshots(p),
            "",  # betrayals — not in PCR struct
            _safe_int(p.get("suicides")),
            best_spree,
            total_time_alive,
        ] + [gt_cols[col] for col in GAMETYPE_STAT_COLUMNS]
        _append_row(ws, row)

    _style_header_row(ws)
    _auto_column_widths(ws, headers)


def _rampant_medal_stats(wb, game):
    """Sheet 5: Medal Stats — 24 individual medal columns."""
    ws = wb.create_sheet("Medal Stats")
    headers = ["player"] + MEDAL_COLUMNS
    ws.append(headers)

    players = _get_valid_players(game)
    sorted_players = sorted(players, key=lambda p: _safe_int(p.get("place")))

    for p in sorted_players:
        bitmask = _player_medals_bitmask(p)
        bits = _decode_medal_bits(bitmask)
        _append_row(ws, [p.get("name", "")] + bits)

    _style_header_row(ws)
    _auto_column_widths(ws, headers)


def _export_rampant(game, output_path):
    """Export a single game in Rampant-compatible format."""
    wb = Workbook()
    _rampant_game_details(wb, game)
    _rampant_post_game_report(wb, game)
    _rampant_versus(wb, game)
    _rampant_game_statistics(wb, game)
    _rampant_medal_stats(wb, game)
    wb.save(output_path)


# ===========================================================================
# Per-game: BUNGIE mode
# ===========================================================================

def _bungie_style_row(ws, row_num, num_cols, fill=None, font=None, alignment=None):
    """Apply styling to an entire row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment


def _bungie_carnage_report(wb, game):
    """Sheet 1: Carnage Report — title block + scoreboard."""
    ws = wb.active
    ws.title = "Carnage Report"

    gametype = (game.get("gametype") or "").lower()
    gametype_display = gametype.upper() if gametype else "UNKNOWN"
    ts = game["_timestamp"]
    players = _get_valid_players(game)
    teams = game.get("teams")
    sorted_players = sorted(players, key=lambda p: _safe_int(p.get("place")))

    # --- Title block ---
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="HALO 2 \u2014 POST-GAME CARNAGE REPORT")
    title_cell.font = BUNGIE_TITLE_FONT
    title_cell.fill = BUNGIE_DARK
    _bungie_style_row(ws, 1, 15, fill=BUNGIE_DARK)

    # Blank row
    _bungie_style_row(ws, 2, 15, fill=BUNGIE_DARK)

    # Game info
    info_rows = [
        ("Gametype", gametype_display),
        ("Date", ts.strftime("%B %d, %Y  %I:%M %p")),
        ("Players", str(len(players))),
    ]
    current_row = 3
    for label, value in info_rows:
        ws.cell(row=current_row, column=1, value=label).font = BUNGIE_LABEL_FONT
        ws.cell(row=current_row, column=2, value=value).font = BUNGIE_VALUE_FONT
        _bungie_style_row(ws, current_row, 15, fill=BUNGIE_DARK)
        current_row += 1

    # Team standings
    if teams:
        _bungie_style_row(ws, current_row, 15, fill=BUNGIE_DARK)
        current_row += 1
        sorted_teams = sorted(teams, key=lambda t: _safe_int(t.get("place")))
        for t in sorted_teams:
            place_str = t.get("place_string", "")
            label = f"{place_str} \u2014 {t.get('name', '')}" if place_str else t.get("name", "")
            ws.cell(row=current_row, column=1, value=label).font = BUNGIE_SUBTITLE_FONT
            ws.cell(row=current_row, column=2, value=_safe_int(t.get("score"))).font = BUNGIE_VALUE_FONT
            _bungie_style_row(ws, current_row, 15, fill=BUNGIE_DARK)
            current_row += 1

    # Blank separator
    _bungie_style_row(ws, current_row, 15, fill=BUNGIE_DARK)
    current_row += 1

    # --- Scoreboard ---
    headers = ["Player", "Score", "Place", "Kills", "Deaths", "Assists", "K/D", "Suicides"]

    # Add accuracy columns if any player has shots
    has_accuracy = any(_player_total_shots(p) > 0 for p in players)
    if has_accuracy:
        headers += ["Shots Fired", "Shots Hit", "Accuracy", "Headshots"]

    # Add medals column if any player has medals
    has_medals = any(_player_medals_total(p) > 0 for p in players)
    if has_medals:
        headers.append("Medals")

    # Add gametype-specific columns
    gt_labels = GAMETYPE_LABELS.get(gametype)
    if gt_labels and any(v != 0 for p in players for v in _gt_values(p)):
        headers += list(gt_labels)

    header_row = current_row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = BUNGIE_HEADER_FONT
        cell.fill = BUNGIE_DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = BUNGIE_THIN_BORDER
    current_row += 1

    # Player rows
    kd_col = headers.index("K/D") + 1
    for i, p in enumerate(sorted_players):
        row_num = current_row
        col = 1

        def _write(value, fmt=None):
            nonlocal col
            cell = ws.cell(row=row_num, column=col, value=_sanitize(value) if isinstance(value, str) else value)
            cell.border = BUNGIE_THIN_BORDER
            cell.alignment = Alignment(horizontal="center") if col > 1 else Alignment()
            if fmt:
                cell.number_format = fmt
            # Alternating row fill
            if i % 2 == 1:
                cell.fill = BUNGIE_ACCENT
            # Gold highlight for 1st place
            if _safe_int(p.get("place")) == 0:
                cell.fill = BUNGIE_GOLD_ROW
            col += 1

        _write(p.get("name", ""))
        _write(p.get("score_string", ""))
        _write(p.get("place_string") or f"#{_safe_int(p.get('place')) + 1}")
        _write(_safe_int(p.get("kills")))
        _write(_safe_int(p.get("deaths")))
        _write(_safe_int(p.get("assists")))
        _write(round(_player_kd(p), 2), "0.00")
        _write(_safe_int(p.get("suicides")))

        if has_accuracy:
            _write(_player_total_shots(p))
            _write(_player_shots_hit(p))
            _write(round(_player_accuracy(p), 1), "0.0")
            _write(_player_headshots(p))

        if has_medals:
            _write(_player_medals_total(p))

        if gt_labels and any(v != 0 for pp in players for v in _gt_values(pp)):
            gt0, gt1 = _gt_values(p)
            _write(gt0)
            _write(gt1)

        current_row += 1

    # Freeze at scoreboard header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(headers[col_idx - 1]))
        ws.column_dimensions[col_letter].width = max(header_len + COL_WIDTH_PADDING, 10)
    # Player name column wider
    ws.column_dimensions["A"].width = 20


def _bungie_versus(wb, game):
    """Sheet 2: Versus — styled NxN kill matrix."""
    ws = wb.create_sheet("Versus")
    players = _get_valid_players(game)
    if not players:
        return

    names = [p.get("name", "") for p in players]
    num_players = len(names)

    # Header row
    header_cell = ws.cell(row=1, column=1, value="Killed \u2193 / By \u2192")
    header_cell.font = BUNGIE_HEADER_FONT
    header_cell.fill = BUNGIE_DARK
    header_cell.border = BUNGIE_THIN_BORDER

    for j, name in enumerate(names):
        cell = ws.cell(row=1, column=j + 2, value=_sanitize(name))
        cell.font = BUNGIE_HEADER_FONT
        cell.fill = BUNGIE_DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = BUNGIE_THIN_BORDER

    # Data rows
    for i, p in enumerate(players):
        row_num = i + 2
        # Row label
        label_cell = ws.cell(row=row_num, column=1, value=_sanitize(names[i]))
        label_cell.font = Font(bold=True, color="c4a747")
        label_cell.fill = BUNGIE_DARK
        label_cell.border = BUNGIE_THIN_BORDER

        killed = _get_killed_array(p)
        for j in range(num_players):
            cell = ws.cell(row=row_num, column=j + 2)
            cell.border = BUNGIE_THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            val = _safe_int(killed[j]) if j < len(killed) else 0

            if i == j:
                # Diagonal — self kills (suicides)
                cell.fill = BUNGIE_GRAY_DIAG
                cell.value = _safe_int(p.get("suicides"))
                cell.font = Font(color="888888")
            elif val > 0:
                cell.value = val
            else:
                cell.value = ""  # blank for zero

    # Column widths
    ws.column_dimensions["A"].width = 20
    for j in range(num_players):
        col_letter = get_column_letter(j + 2)
        ws.column_dimensions[col_letter].width = max(len(names[j]) + 2, 8)


def _bungie_medals(wb, game):
    """Sheet 3: Medals — only medals earned in this game."""
    players = _get_valid_players(game)

    # Collect all medal bits across all players
    combined_bitmask = 0
    for p in players:
        combined_bitmask |= _player_medals_bitmask(p)

    if combined_bitmask == 0:
        return  # No medals — skip sheet entirely

    # Find which medal indices are present
    active_indices = [i for i in range(24) if combined_bitmask & (1 << i)]
    active_names = [MEDAL_DISPLAY_NAMES[i] for i in active_indices]

    ws = wb.create_sheet("Medals")
    headers = ["Player"] + active_names
    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = BUNGIE_HEADER_FONT
        cell.fill = BUNGIE_DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = BUNGIE_THIN_BORDER
    ws.freeze_panes = "A2"

    sorted_players = sorted(players, key=lambda p: _safe_int(p.get("place")))
    for row_idx, p in enumerate(sorted_players, 2):
        bitmask = _player_medals_bitmask(p)
        ws.cell(row=row_idx, column=1, value=_sanitize(p.get("name", ""))).border = BUNGIE_THIN_BORDER
        for col_offset, medal_idx in enumerate(active_indices):
            cell = ws.cell(row=row_idx, column=col_offset + 2)
            val = 1 if bitmask & (1 << medal_idx) else 0
            cell.value = val if val else ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = BUNGIE_THIN_BORDER

    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = len(headers[col_idx - 1]) + 3


def _bungie_accuracy(wb, game):
    """Sheet 4: Accuracy — detailed shooting breakdown."""
    players = _get_valid_players(game)
    shooters = [p for p in players if _player_total_shots(p) > 0]
    if not shooters:
        return  # No one fired — skip sheet

    ws = wb.create_sheet("Accuracy")
    headers = ["Player", "Shots Fired", "Shots Hit", "Accuracy", "Headshots", "HS Rate"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = BUNGIE_HEADER_FONT
        cell.fill = BUNGIE_DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = BUNGIE_THIN_BORDER
    ws.freeze_panes = "A2"

    # Sort by accuracy descending
    shooters.sort(key=lambda p: _player_accuracy(p), reverse=True)

    for row_idx, p in enumerate(shooters, 2):
        total = _player_total_shots(p)
        hit = _player_shots_hit(p)
        hs = _player_headshots(p)
        acc = round(_player_accuracy(p), 1)
        hs_rate = round(hs / max(hit, 1) * 100, 1)

        ws.cell(row=row_idx, column=1, value=_sanitize(p.get("name", ""))).border = BUNGIE_THIN_BORDER
        for col, val in enumerate([total, hit, acc, hs, hs_rate], 2):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BUNGIE_THIN_BORDER
            if col in (4, 6):
                cell.number_format = "0.0"

    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = len(headers[col_idx - 1]) + 4


def _export_bungie(game, output_path):
    """Export a single game in Bungie Studios style."""
    wb = Workbook()
    _bungie_carnage_report(wb, game)
    _bungie_versus(wb, game)
    _bungie_medals(wb, game)
    _bungie_accuracy(wb, game)
    wb.save(output_path)


# ===========================================================================
# Per-game export orchestrator
# ===========================================================================

def _export_per_game(games, output_dir, style="bungie"):
    """Export each game as a separate .xlsx file."""
    os.makedirs(output_dir, exist_ok=True)

    exported = 0
    skipped = 0

    for game in games:
        if not _game_has_valid_players(game):
            skipped += 1
            continue

        ts = game["_timestamp"]
        gametype = game.get("gametype") or "unknown"
        filename = f"{ts.strftime('%Y-%m-%d_%H-%M-%S')}_{gametype}.xlsx"
        output_path = os.path.join(output_dir, filename)

        if style == "rampant":
            _export_rampant(game, output_path)
        else:
            _export_bungie(game, output_path)

        exported += 1
        print(f"  {filename}")

    return exported, skipped


# ===========================================================================
# PGCR mode — one workbook, one sheet per game, in-game PGCR layout
# ===========================================================================

# Additional styling for PGCR section headers
PGCR_SECTION_FONT = Font(bold=True, size=12, color="c4a747")
PGCR_INFO_FONT = Font(size=11, color="AAAAAA")

# Team colors for PGCR
TEAM_RED_FILL = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")  # Red
TEAM_BLUE_FILL = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Blue
GREY_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")  # Grey
TEAM_FONT = Font(color="FFFFFF", bold=True)  # White text for contrast


def _pgcr_sheet_name(game, used_names):
    """Generate a unique sheet name like 'Slayer 02-17' for a game."""
    ts = game["_timestamp"]
    gametype = (game.get("gametype") or "Unknown").capitalize()
    base = f"{gametype} {ts.strftime('%m-%d')}"
    name = base[:31]  # Excel sheet name limit is 31 chars
    # Ensure uniqueness
    counter = 2
    while name in used_names:
        suffix = f" ({counter})"
        name = (base[:31 - len(suffix)] + suffix)
        counter += 1
    used_names.add(name)
    return name


def _DEAD_START(wb, game):
    num_cols = 15

    # --- TITLE BLOCK ---
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="HALO 2 \u2014 POST-GAME CARNAGE REPORT")
    title_cell.font = BUNGIE_TITLE_FONT
    title_cell.fill = BUNGIE_DARK
    _pgcr_fill_dark_row(ws, 1, num_cols)

    # Game info
    _pgcr_fill_dark_row(ws, 2, num_cols)
    info_text = f"Gametype: {gametype_display}  |  Date: {ts.strftime('%B %d, %Y  %I:%M %p')}  |  Players: {len(players)}"
    info_cell = ws.cell(row=2, column=1, value=info_text)
    info_cell.font = PGCR_INFO_FONT
    info_cell.fill = BUNGIE_DARK

    current_row = 3
    _pgcr_fill_dark_row(ws, current_row, num_cols)
    current_row += 1

    # --- TEAM STANDINGS ---
    if teams:
        current_row = _pgcr_write_section_header(ws, current_row, "TEAM STANDINGS", num_cols)
        team_headers = ["Team", "Place", "Score"]
        current_row = _pgcr_write_column_headers(ws, current_row, team_headers, num_cols)

        sorted_teams = sorted(teams, key=lambda t: _safe_int(t.get("place")))
        for i, t in enumerate(sorted_teams):
            is_first = _safe_int(t.get("place")) == 0
            _pgcr_write_data_cell(ws, current_row, 1, t.get("name", ""), is_first, i)
            _pgcr_write_data_cell(ws, current_row, 2, t.get("place_string", ""), is_first, i)
            _pgcr_write_data_cell(ws, current_row, 3, t.get("score_string", str(_safe_int(t.get("score")))), is_first, i)
            _pgcr_fill_dark_row(ws, current_row, num_cols)
            # Re-style the data cells (fill_dark_row overwrites fill)
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                if is_first:
                    cell.fill = BUNGIE_GOLD_ROW
                elif i % 2 == 1:
                    cell.fill = BUNGIE_ACCENT
            current_row += 1

        _pgcr_fill_dark_row(ws, current_row, num_cols)
        current_row += 1

    # --- SCORE (gametype-specific player stats) ---
    current_row = _pgcr_write_section_header(ws, current_row, "SCORE", num_cols)

    score_headers = ["Player", "Place"]
    gt_labels = GAMETYPE_LABELS.get(gametype)
    has_gt = gt_labels and any(v != 0 for p in players for v in _gt_values(p))
    if has_gt:
        score_headers.extend(gt_labels)
    score_headers.append("Score")
    current_row = _pgcr_write_column_headers(ws, current_row, score_headers, num_cols)

    for i, p in enumerate(sorted_players):
        is_first = _safe_int(p.get("place")) == 0
        col = 1
        _pgcr_write_data_cell(ws, current_row, col, p.get("name", ""), is_first, i); col += 1
        _pgcr_write_data_cell(ws, current_row, col, p.get("place_string") or f"#{_safe_int(p.get('place')) + 1}", is_first, i); col += 1
        if has_gt:
            gt0, gt1 = _gt_values(p)
            _pgcr_write_data_cell(ws, current_row, col, gt0, is_first, i); col += 1
            _pgcr_write_data_cell(ws, current_row, col, gt1, is_first, i); col += 1
        _pgcr_write_data_cell(ws, current_row, col, p.get("score_string", ""), is_first, i)
        _pgcr_fill_dark_row(ws, current_row, num_cols)
        # Restore data cell styling after fill_dark_row
        for c in range(1, col + 1):
            cell = ws.cell(row=current_row, column=c)
            if is_first:
                cell.fill = BUNGIE_GOLD_ROW
            elif i % 2 == 1:
                cell.fill = BUNGIE_ACCENT
        current_row += 1

    _pgcr_fill_dark_row(ws, current_row, num_cols)
    current_row += 1

    # --- KILLS ---
    current_row = _pgcr_write_section_header(ws, current_row, "KILLS", num_cols)
    kill_headers = ["Player", "Kills", "Assists", "Deaths", "Suicides"]
    current_row = _pgcr_write_column_headers(ws, current_row, kill_headers, num_cols)

    for i, p in enumerate(sorted_players):
        is_first = _safe_int(p.get("place")) == 0
        vals = [
            p.get("name", ""),
            _safe_int(p.get("kills")),
            _safe_int(p.get("assists")),
            _safe_int(p.get("deaths")),
            _safe_int(p.get("suicides")),
        ]
        for col, v in enumerate(vals, 1):
            _pgcr_write_data_cell(ws, current_row, col, v, is_first, i)
        _pgcr_fill_dark_row(ws, current_row, num_cols)
        for c in range(1, len(vals) + 1):
            cell = ws.cell(row=current_row, column=c)
            if is_first:
                cell.fill = BUNGIE_GOLD_ROW
            elif i % 2 == 1:
                cell.fill = BUNGIE_ACCENT
        current_row += 1

    _pgcr_fill_dark_row(ws, current_row, num_cols)
    current_row += 1

    # --- HIT STATS ---
    has_accuracy = any(_player_total_shots(p) > 0 for p in players)
    if has_accuracy:
        current_row = _pgcr_write_section_header(ws, current_row, "HIT STATS", num_cols)
        hit_headers = ["Player", "Shots Fired", "Shots Hit", "Hit %", "Headshots"]
        current_row = _pgcr_write_column_headers(ws, current_row, hit_headers, num_cols)

        for i, p in enumerate(sorted_players):
            is_first = _safe_int(p.get("place")) == 0
            acc = round(_player_accuracy(p), 1)
            vals = [
                p.get("name", ""),
                _player_total_shots(p),
                _player_shots_hit(p),
                acc,
                _player_headshots(p),
            ]
            for col, v in enumerate(vals, 1):
                fmt = "0.0" if col == 4 else None
                _pgcr_write_data_cell(ws, current_row, col, v, is_first, i, fmt=fmt)
            _pgcr_fill_dark_row(ws, current_row, num_cols)
            for c in range(1, len(vals) + 1):
                cell = ws.cell(row=current_row, column=c)
                if is_first:
                    cell.fill = BUNGIE_GOLD_ROW
                elif i % 2 == 1:
                    cell.fill = BUNGIE_ACCENT
            current_row += 1

        _pgcr_fill_dark_row(ws, current_row, num_cols)
        current_row += 1

    # --- MEDALS ---
    combined_bitmask = 0
    for p in players:
        combined_bitmask |= _player_medals_bitmask(p)

    if combined_bitmask != 0:
        current_row = _pgcr_write_section_header(ws, current_row, "MEDALS", num_cols)
        active_indices = [i for i in range(24) if combined_bitmask & (1 << i)]
        active_names = [MEDAL_DISPLAY_NAMES[i] for i in active_indices]
        medal_headers = ["Player", "Total"] + active_names
        current_row = _pgcr_write_column_headers(ws, current_row, medal_headers, num_cols)

        for i, p in enumerate(sorted_players):
            is_first = _safe_int(p.get("place")) == 0
            bitmask = _player_medals_bitmask(p)
            _pgcr_write_data_cell(ws, current_row, 1, p.get("name", ""), is_first, i)
            _pgcr_write_data_cell(ws, current_row, 2, _player_medals_total(p), is_first, i)
            for col_offset, medal_idx in enumerate(active_indices):
                val = 1 if bitmask & (1 << medal_idx) else ""
                _pgcr_write_data_cell(ws, current_row, col_offset + 3, val, is_first, i)
            total_medal_cols = len(medal_headers)
            _pgcr_fill_dark_row(ws, current_row, num_cols)
            for c in range(1, total_medal_cols + 1):
                cell = ws.cell(row=current_row, column=c)
                if is_first:
                    cell.fill = BUNGIE_GOLD_ROW
                elif i % 2 == 1:
                    cell.fill = BUNGIE_ACCENT
            current_row += 1

        _pgcr_fill_dark_row(ws, current_row, num_cols)
        current_row += 1

    # --- PLAYER VS. PLAYER ---
    if len(players) > 1:
        current_row = _pgcr_write_section_header(ws, current_row, "PLAYER VS. PLAYER", num_cols)

        names = [p.get("name", "") for p in sorted_players]

        # Header row: "Killed / By" + player names
        header_cell = ws.cell(row=current_row, column=1, value="Killed \u2193 / By \u2192")
        header_cell.font = BUNGIE_HEADER_FONT
        header_cell.fill = BUNGIE_DARK
        header_cell.border = BUNGIE_THIN_BORDER
        for j, name in enumerate(names):
            cell = ws.cell(row=current_row, column=j + 2, value=_sanitize(name))
            cell.font = BUNGIE_HEADER_FONT
            cell.fill = BUNGIE_DARK
            cell.alignment = Alignment(horizontal="center")
            cell.border = BUNGIE_THIN_BORDER
        _pgcr_fill_dark_row(ws, current_row, num_cols)
        # Restore header styling after fill
        ws.cell(row=current_row, column=1).font = BUNGIE_HEADER_FONT
        ws.cell(row=current_row, column=1).fill = BUNGIE_DARK
        ws.cell(row=current_row, column=1).border = BUNGIE_THIN_BORDER
        for j in range(len(names)):
            cell = ws.cell(row=current_row, column=j + 2)
            cell.font = BUNGIE_HEADER_FONT
            cell.fill = BUNGIE_DARK
            cell.border = BUNGIE_THIN_BORDER
        current_row += 1

        # Data rows
        # Build slot-index mapping: sorted_players may differ from slot order
        # killed[] is indexed by slot position, so we need to map sorted player indices
        # to their original slot positions
        slot_indices = []
        all_players = game.get("players", [])
        for sp in sorted_players:
            sp_name = sp.get("name", "")
            for slot_idx, ap in enumerate(all_players):
                if ap.get("name", "") == sp_name:
                    slot_indices.append(slot_idx)
                    break
            else:
                slot_indices.append(0)

        for i, p in enumerate(sorted_players):
            row_num = current_row
            # Row label
            label_cell = ws.cell(row=row_num, column=1, value=_sanitize(names[i]))
            label_cell.font = Font(bold=True, color="c4a747")
            label_cell.fill = BUNGIE_DARK
            label_cell.border = BUNGIE_THIN_BORDER

            killed = _get_killed_array(p)
            for j in range(len(sorted_players)):
                cell = ws.cell(row=row_num, column=j + 2)
                cell.border = BUNGIE_THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                target_slot = slot_indices[j]

                if i == j:
                    # Diagonal — suicides
                    cell.fill = BUNGIE_GRAY_DIAG
                    cell.value = _safe_int(p.get("suicides"))
                    cell.font = Font(color="888888")
                else:
                    val = _safe_int(killed[target_slot]) if target_slot < len(killed) else 0
                    cell.value = val if val > 0 else ""
                    cell.font = BUNGIE_VALUE_FONT

            _pgcr_fill_dark_row(ws, row_num, num_cols)
            # Restore PvP data cells
            ws.cell(row=row_num, column=1).font = Font(bold=True, color="c4a747")
            ws.cell(row=row_num, column=1).fill = BUNGIE_DARK
            ws.cell(row=row_num, column=1).border = BUNGIE_THIN_BORDER
            for j in range(len(sorted_players)):
                cell = ws.cell(row=row_num, column=j + 2)
                cell.border = BUNGIE_THIN_BORDER
                if i == j:
                    cell.fill = BUNGIE_GRAY_DIAG
            current_row += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Freeze at row 3 (below title + info)
    ws.freeze_panes = "A3"

    return ws


def _pgcr_player_totals_sheet(wb, games):
    """Build Player Totals sheet with aggregate stats ranked best to worst."""
    ws = wb.create_sheet("Player Totals")

    # Aggregate stats by player
    agg = {}
    for g in games:
        for p in g.get("players", []):
            name = p.get("name", "")
            if not name or not _is_valid_name(name):
                continue
            if name not in agg:
                agg[name] = {
                    "games": 0, "kills": 0, "deaths": 0, "assists": 0,
                    "suicides": 0, "medals": 0, "headshots": 0,
                    "total_shots": 0, "shots_hit": 0,
                }
            s = agg[name]
            s["games"] += 1
            s["kills"] += _safe_int(p.get("kills"))
            s["deaths"] += _safe_int(p.get("deaths"))
            s["assists"] += _safe_int(p.get("assists"))
            s["suicides"] += _safe_int(p.get("suicides"))
            s["medals"] += _player_medals_total(p)
            s["headshots"] += _player_headshots(p)
            s["total_shots"] += _player_total_shots(p)
            s["shots_hit"] += _player_shots_hit(p)

    # Sort by kills descending
    sorted_players = sorted(agg.items(), key=lambda kv: kv[1]["kills"], reverse=True)

    # --- TITLE ---
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="PLAYER TOTALS — ALL GAMES")
    title_cell.font = BUNGIE_TITLE_FONT
    title_cell.fill = BUNGIE_DARK
    for col in range(1, 9):
        ws.cell(row=1, column=col).fill = BUNGIE_DARK

    # Blank row
    for col in range(1, 9):
        ws.cell(row=2, column=col).fill = BUNGIE_DARK

    # Headers
    headers = ["Player", "Games", "Kills", "Deaths", "Assists", "K/D", "Accuracy%", "Headshots", "Medals"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = BUNGIE_HEADER_FONT
        cell.fill = BUNGIE_DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = BUNGIE_THIN_BORDER
    for col in range(1, 10):
        ws.cell(row=3, column=col).fill = BUNGIE_DARK

    # Data rows (ranked by kills) - grey only
    for i, (name, s) in enumerate(sorted_players):
        row_num = i + 4
        kd = s["kills"] / max(s["deaths"], 1)
        acc = (s["shots_hit"] / s["total_shots"] * 100) if s["total_shots"] > 0 else 0.0

        vals = [
            name,
            s["games"],
            s["kills"],
            s["deaths"],
            s["assists"],
            round(kd, 2),
            round(acc, 1),
            s["headshots"],
            s["medals"],
        ]

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=_sanitize(v) if isinstance(v, str) else v)
            cell.border = BUNGIE_THIN_BORDER
            cell.alignment = Alignment(horizontal="center") if col > 1 else Alignment()
            cell.font = BUNGIE_VALUE_FONT
            cell.fill = GREY_FILL
        # Fill remaining cols grey
        for col in range(len(vals) + 1, 10):
            c = ws.cell(row=row_num, column=col)
            c.fill = GREY_FILL

    # Column widths
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "A4"


def _pgcr_index_sheet(wb, games):
    """Build the Index sheet listing all games."""
    ws = wb.active
    ws.title = "Index"
    headers = ["Game #", "Date", "Time", "Gametype", "Players", "Result"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = BUNGIE_HEADER_FONT
        cell.fill = BUNGIE_DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = BUNGIE_THIN_BORDER
    ws.freeze_panes = "A2"

    for game_num, g in enumerate(games, 1):
        if not _game_has_valid_players(g):
            continue
        ts = g["_timestamp"]
        gametype = (g.get("gametype") or "").capitalize()
        players = _get_valid_players(g)
        # Determine result (1st place player or team)
        teams = g.get("teams")
        if teams:
            sorted_teams = sorted(teams, key=lambda t: _safe_int(t.get("place")))
            result = f"{sorted_teams[0].get('name', '')} wins" if sorted_teams else ""
        else:
            sorted_p = sorted(players, key=lambda p: _safe_int(p.get("place")))
            result = f"{sorted_p[0].get('name', '')} wins" if sorted_p else ""

        row_num = game_num + 1
        vals = [game_num, ts.strftime("%Y-%m-%d"), ts.strftime("%H:%M:%S"),
                gametype, len(players), result]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=_sanitize(v) if isinstance(v, str) else v)
            cell.font = BUNGIE_VALUE_FONT
            cell.fill = BUNGIE_DARK
            cell.border = BUNGIE_THIN_BORDER
            cell.alignment = Alignment(horizontal="center") if col > 1 else Alignment()

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = BUNGIE_DARK
        ws.cell(row=1, column=col).font = BUNGIE_HEADER_FONT

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 24


def _export_pgcr(games, output_path):
    """Export all games to a single workbook in PGCR format (one sheet per game)."""
    wb = Workbook()
    _pgcr_index_sheet(wb, games)
    _pgcr_player_totals_sheet(wb, games)

    used_names = {"Index", "Player Totals"}
    exported = 0
    for game in games:
        if not _game_has_valid_players(game):
            continue
        sheet_name = _pgcr_sheet_name(game, used_names)
        # Create the sheet and make it the target
        ws = wb.create_sheet(title=sheet_name)
        # We need to build content on this specific sheet, so temporarily
        # set it as active for _pgcr_game_sheet to use
        game["_sheet_name"] = sheet_name
        _pgcr_game_sheet_on(wb, ws, game)
        exported += 1

    wb.save(output_path)
    return exported


def _pgcr_game_sheet_on(wb, ws, game):
    """Build PGCR content on a specific worksheet (not using wb.active)."""
    gametype = (game.get("gametype") or "").lower()
    gametype_display = gametype.upper() if gametype else "UNKNOWN"
    ts = game["_timestamp"]
    players = _get_valid_players(game)
    teams = game.get("teams")
    sorted_players = sorted(players, key=lambda p: _safe_int(p.get("place")))
    num_cols = 15

    # --- TITLE BLOCK ---
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="HALO 2 \u2014 POST-GAME CARNAGE REPORT")
    title_cell.font = BUNGIE_TITLE_FONT
    title_cell.fill = BUNGIE_DARK
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col).fill = BUNGIE_DARK

    # Game info
    for col in range(1, num_cols + 1):
        ws.cell(row=2, column=col).fill = BUNGIE_DARK
    info_text = f"Gametype: {gametype_display}  |  Date: {ts.strftime('%B %d, %Y  %I:%M %p')}  |  Players: {len(players)}"
    info_cell = ws.cell(row=2, column=1, value=info_text)
    info_cell.font = PGCR_INFO_FONT
    info_cell.fill = BUNGIE_DARK

    current_row = 3
    for col in range(1, num_cols + 1):
        ws.cell(row=current_row, column=col).fill = BUNGIE_DARK
    current_row += 1

    def _fill_dark(row):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = BUNGIE_DARK

    def _section_header(row, title):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = PGCR_SECTION_FONT
        cell.fill = BUNGIE_DARK
        _fill_dark(row)
        return row + 1

    def _col_headers(row, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = BUNGIE_HEADER_FONT
            cell.fill = BUNGIE_DARK
            cell.alignment = Alignment(horizontal="center")
            cell.border = BUNGIE_THIN_BORDER
        _fill_dark(row)
        # Re-apply header styling after fill
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = BUNGIE_HEADER_FONT
            cell.fill = BUNGIE_DARK
            cell.border = BUNGIE_THIN_BORDER
        return row + 1

    def _data_row(row, values, is_first, row_idx, fmts=None, team=None):
        # Choose team color: red for team 0, blue for team 1, grey otherwise
        if team == 0:
            fill = TEAM_RED_FILL
            font = TEAM_FONT
        elif team == 1:
            fill = TEAM_BLUE_FILL
            font = TEAM_FONT
        else:
            fill = GREY_FILL
            font = BUNGIE_VALUE_FONT

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=_sanitize(v) if isinstance(v, str) else v)
            cell.border = BUNGIE_THIN_BORDER
            cell.alignment = Alignment(horizontal="center") if col > 1 else Alignment()
            cell.font = font
            if fmts and col in fmts:
                cell.number_format = fmts[col]
            cell.fill = fill
        # Fill remaining cols
        for col in range(len(values) + 1, num_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill
        return row + 1

    def _blank_dark(row):
        _fill_dark(row)
        return row + 1

    # --- TEAM STANDINGS ---
    if teams:
        current_row = _section_header(current_row, "TEAM STANDINGS")
        current_row = _col_headers(current_row, ["Team", "Place", "Score"])
        sorted_teams = sorted(teams, key=lambda t: _safe_int(t.get("place")))
        for i, t in enumerate(sorted_teams):
            is_first = _safe_int(t.get("place")) == 0
            current_row = _data_row(current_row, [
                t.get("name", ""),
                t.get("place_string", ""),
                t.get("score_string", str(_safe_int(t.get("score")))),
            ], is_first, i)
        current_row = _blank_dark(current_row)

    # --- SCORE ---
    current_row = _section_header(current_row, "SCORE")
    score_headers = ["Player", "Place"]
    gt_labels = GAMETYPE_LABELS.get(gametype)
    has_gt = gt_labels and any(v != 0 for p in players for v in _gt_values(p))
    if has_gt:
        score_headers.extend(gt_labels)
    score_headers.append("Score")
    current_row = _col_headers(current_row, score_headers)

    for i, p in enumerate(sorted_players):
        is_first = _safe_int(p.get("place")) == 0
        team = _safe_int(p.get("team"))
        vals = [p.get("name", ""), p.get("place_string") or f"#{_safe_int(p.get('place')) + 1}"]
        if has_gt:
            gt0, gt1 = _gt_values(p)
            vals.extend([gt0, gt1])
        vals.append(p.get("score_string", ""))
        current_row = _data_row(current_row, vals, is_first, i, team=team)
    current_row = _blank_dark(current_row)

    # --- KILLS ---
    current_row = _section_header(current_row, "KILLS")
    current_row = _col_headers(current_row, ["Player", "Kills", "Assists", "Deaths", "Suicides"])
    for i, p in enumerate(sorted_players):
        is_first = _safe_int(p.get("place")) == 0
        team = _safe_int(p.get("team"))
        current_row = _data_row(current_row, [
            p.get("name", ""), _safe_int(p.get("kills")), _safe_int(p.get("assists")),
            _safe_int(p.get("deaths")), _safe_int(p.get("suicides")),
        ], is_first, i, team=team)
    current_row = _blank_dark(current_row)

    # --- HIT STATS ---
    if any(_player_total_shots(p) > 0 for p in players):
        current_row = _section_header(current_row, "HIT STATS")
        current_row = _col_headers(current_row, ["Player", "Shots Fired", "Shots Hit", "Hit %", "Headshots"])
        for i, p in enumerate(sorted_players):
            is_first = _safe_int(p.get("place")) == 0
            team = _safe_int(p.get("team"))
            current_row = _data_row(current_row, [
                p.get("name", ""), _player_total_shots(p), _player_shots_hit(p),
                round(_player_accuracy(p), 1), _player_headshots(p),
            ], is_first, i, fmts={4: "0.0"}, team=team)
        current_row = _blank_dark(current_row)

    # --- MEDALS ---
    combined_bitmask = 0
    for p in players:
        combined_bitmask |= _player_medals_bitmask(p)

    if combined_bitmask != 0:
        current_row = _section_header(current_row, "MEDALS")
        active_indices = [i for i in range(24) if combined_bitmask & (1 << i)]
        active_names = [MEDAL_DISPLAY_NAMES[i] for i in active_indices]
        medal_headers = ["Player", "Total"] + active_names
        current_row = _col_headers(current_row, medal_headers)

        for i, p in enumerate(sorted_players):
            is_first = _safe_int(p.get("place")) == 0
            team = _safe_int(p.get("team"))
            bitmask = _player_medals_bitmask(p)
            vals = [p.get("name", ""), _player_medals_total(p)]
            for medal_idx in active_indices:
                vals.append(1 if bitmask & (1 << medal_idx) else "")
            current_row = _data_row(current_row, vals, is_first, i, team=team)
        current_row = _blank_dark(current_row)

    # --- PLAYER VS. PLAYER ---
    if len(sorted_players) > 1:
        current_row = _section_header(current_row, "PLAYER VS. PLAYER")
        names = [p.get("name", "") for p in sorted_players]

        # Map sorted player indices to their slot positions in the original array
        all_players = game.get("players", [])
        slot_indices = []
        for sp in sorted_players:
            sp_name = sp.get("name", "")
            for slot_idx, ap in enumerate(all_players):
                if ap.get("name", "") == sp_name:
                    slot_indices.append(slot_idx)
                    break
            else:
                slot_indices.append(0)

        # Header row
        pvp_headers = ["Killed \u2193 / By \u2192"] + [_sanitize(n) for n in names]
        current_row = _col_headers(current_row, pvp_headers)

        # Data rows
        for i, p in enumerate(sorted_players):
            row_num = current_row
            team = _safe_int(p.get("team"))

            # Choose team color for row label
            if team == 0:
                label_fill = TEAM_RED_FILL
                label_font = TEAM_FONT
            elif team == 1:
                label_fill = TEAM_BLUE_FILL
                label_font = TEAM_FONT
            else:
                label_fill = GREY_FILL
                label_font = BUNGIE_VALUE_FONT

            # Row label with team color
            label_cell = ws.cell(row=row_num, column=1, value=_sanitize(names[i]))
            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.border = BUNGIE_THIN_BORDER

            killed = _get_killed_array(p)
            for j in range(len(sorted_players)):
                cell = ws.cell(row=row_num, column=j + 2)
                cell.border = BUNGIE_THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                target_slot = slot_indices[j]

                if i == j:
                    cell.fill = BUNGIE_GRAY_DIAG
                    cell.value = _safe_int(p.get("suicides"))
                    cell.font = Font(color="888888")
                else:
                    val = _safe_int(killed[target_slot]) if target_slot < len(killed) else 0
                    cell.value = val if val > 0 else ""
                    cell.font = BUNGIE_VALUE_FONT
                    cell.fill = label_fill

            # Fill remaining columns with team color
            for col in range(len(sorted_players) + 2, num_cols + 1):
                ws.cell(row=row_num, column=col).fill = label_fill
            current_row += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "A3"


# ===========================================================================
# Main
# ===========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Export Halo 2 game history to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s -o stats.xlsx                                  # All-in-one summary
  %(prog)s --pgcr -o pgcr_report.xlsx                     # PGCR format (one sheet per game)
  %(prog)s --per-game --output-dir exports/               # Per-game (Bungie style)
  %(prog)s --per-game --style rampant --output-dir exports/  # Per-game (Rampant format)
  %(prog)s --per-game --output-dir exports/ --from 2026-02-10  # Date filter
        """,
    )
    parser.add_argument("--history-dir", default="history",
                        help="History directory (default: history/)")
    parser.add_argument("--output", "-o", default="halo2_stats.xlsx",
                        help="Output XLSX file (default: halo2_stats.xlsx)")
    parser.add_argument("--from", dest="date_from",
                        help="Start date filter (YYYY-MM-DD)")
    parser.add_argument("--to", dest="date_to",
                        help="End date filter (YYYY-MM-DD)")
    parser.add_argument("--pgcr", action="store_true",
                        help="PGCR format: one sheet per game matching in-game carnage report layout")
    parser.add_argument("--per-game", action="store_true",
                        help="Export one .xlsx per game instead of all-in-one")
    parser.add_argument("--style", choices=["bungie", "rampant"], default="bungie",
                        help="Per-game export style (default: bungie)")
    parser.add_argument("--output-dir", default="exports",
                        help="Output directory for per-game mode (default: exports/)")
    args = parser.parse_args()

    games = _load_games(args.history_dir, args.date_from, args.date_to)
    if not games:
        print("No games found.", file=sys.stderr)
        sys.exit(1)

    if args.pgcr:
        print(f"Exporting {len(games)} games in PGCR format to {args.output}")
        exported = _export_pgcr(games, args.output)
        print(f"Exported {exported} games to {args.output}")
    elif args.per_game:
        print(f"Exporting {len(games)} games ({args.style} style) to {args.output_dir}/")
        exported, skipped = _export_per_game(games, args.output_dir, args.style)
        print(f"\nExported {exported} games" + (f", skipped {skipped} (invalid data)" if skipped else ""))
    else:
        total_player_rows = sum(len(g.get("players", [])) for g in games)
        wb = Workbook()
        _build_games_sheet(wb, games)
        _build_players_sheet(wb, games)
        _build_summary_sheet(wb, games)
        _build_teams_sheet(wb, games)
        wb.save(args.output)
        print(f"Exported {len(games)} games ({total_player_rows} players) to {args.output}")


if __name__ == "__main__":
    main()
